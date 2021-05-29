"""
author: miaokela
Date: 2021-04-16 13:01:08
LastEditTime: 2021-05-01 14:29:31
Description: 每日采集数量、推送数量统计/指定时间段的采集数量、推送数量 以excel的形式输出
"""
import pymysql
import re
import gevent
import os

from datetime import datetime, timedelta
from openpyxl import Workbook, styles
from openpyxl.styles import Border, Side


class DBQuery(object):
    """
    数据库操作
        MySQL
    """

    __instance = None

    def __new__(cls, **kwargs):
        if not cls.__instance:
            cls.__instance = object.__new__(cls)
        return cls.__instance

    def __del__(self):
        self.con.close()

    def __init__(self, **kwargs):
        self.db_name = kwargs.get('db')
        self.con = pymysql.connect(
            host=kwargs.get('host'),
            user=kwargs.get('user'),
            password=kwargs.get('password'),
            db=self.db_name,
            charset='utf8mb4',
            port=8050
        )
        self.msg = ''

    def fetch_all(self, sql):
        ret = []
        if self.con:
            try:
                with self.con.cursor() as cursor:
                    cursor.execute(sql)
                    ret = cursor.fetchall()
            except Exception as e:
                self.msg = '数据库链接失败: {0}'.format(e)
        else:
            self.msg = '查询SQL({0})时链接断开.'.format(sql)
        return ret

    def exec_sql(self, sql):
        ret = True
        try:
            with self.con.cursor() as cursor:
                cursor.execute(sql)
            self.con.commit()
        except Exception as e:
            self.msg = '{0}执行失败: {1}'.format(sql, e)
            ret = False
        return ret


class ReportOutput(DBQuery):
    """
    采集信息输出
    """
    area_map = {
        '00': '全国公共资源交易平台',
        '01': '中国招标投标公共服务平台',
        '02': '北京市公共资源交易服务平台',
        '03': '天津市公共资源交易网',
        '04': '河北省公共资源交易平台',
        '05': '山西省公共资源交易平台',
        '06': '内蒙古公共资源交易平台',
        '07': '辽宁省公共资源交易平台',
        '08': '吉林公共资源交易平台',
        '10': '黑龙江公共资源交易平台',
        '11': '上海市公共资源交易服务平台',
        '12': '上海建设工程交易服务中心',
        '13': '江苏省公共资源交易服务平台',
        '14': '浙江省公共资源交易服务平台1',
        '15': '浙江省公共资源交易服务平台2',
        '16': '安徽省公共资源交易服务平台',
        '17': '福建省公共资源交易电子公共服务平台',
        '18': '福建省公共资源交易服务平台',
        '19': '江西省公共资源交易服务平台',
        # '20': '江西省公共资源交易服务平台',  # 访问不了
        '21': '山东省公共资源交易服务平台',
        '22': '山东省公共资源交易中心',
        '23': '河南省公共资源交易服务平台',
        '24': '河南省公共资源交易中心门户网',
        '25': '招投标信息_河南省水利厅',
        '26': '湖北省公共资源交易服务平台',
        '27': '湖北省公共资源电子交易服务系统',
        '28': '湖南省公共资源交易服务平台',
        '30': '广东公共资源交易平台',
        '33': '广西省公共资源交易服务平台',
        '36': '甘肃省公共资源交易服务平台',
        '39': '重庆省公共资源交易服务平台',
        '40': '四川省公共资源交易服务平台',
        '41': '贵州省公共资源交易服务平台',
        '42': '云南省公共资源交易服务平台',
        '44': '西藏自治区公共资源交易服务平台',
        '45': '陕西省公共资源交易服务平台',
        '47': '青海省公共资源交易服务平台',
        '49': '宁夏回族自治区公共资源交易网',
        '50': '新疆维吾尔族自治区公共资源交易服务平台',
        '51': '新疆生产建设兵团公共资源交易服务平台',
        '52': '嗨招电子招标采购平台',
        '53': '必联网',
        '55': '天工开物',
        '56': '河北建投集团电子招投标交易平台',
        '54': 'E共享交易平台',
        '57': '精彩纵横',
        '62': '新点电子交易平台',
        '71': '招财进宝',
        '3301': '杭州市公共资源交易平台',
        '3303': '浙能集团电子招标投标交易平台',
        '3304': '浙江省水利厅',
        '3305': '宁波公共资源交易平台',
        '3306': '嘉兴公共资源交易平台',
        '3307': '湖州公共资源交易平台',
        '3308': '衢州公共资源交易平台',
        '3309': '温州公共资源交易平台',
        '3310': '台州公共资源交易平台',
        '3311': '丽水公共资源交易平台',
        '3312': '绍兴公共资源交易平台',
        '3313': '舟山公共资源交易平台',
        '3314': '余杭公共资源交易平台',
        '3315': '柯桥公共资源交易平台',
        '3318': '金华市公共资源交易中心',
        '3319': '长兴公共资源交易平台',
        '3320': '苍南公共资源交易平台',
        '3321': '浙江临海市公告资源交易中心',
        '3322': '安吉公共资源交易中心',
        '3323': '萧山政府门户网站',
        '3324': '南浔区公共资源交易平台',
        '3326': '龙游政府门户网站',
        '3327': '平阳公共资源交易平台',
        '3328': '常山公共资源交易平台',

    }
    border_type = Side(border_style="medium", color='FF000000')
    border = Border(left=border_type,
                    right=border_type,
                    top=border_type,
                    bottom=border_type,
                    diagonal=border_type,
                    diagonal_direction=0,
                    outline=border_type,
                    vertical=border_type,
                    horizontal=border_type
                    )

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.sdt = kwargs.get('sdt')
        self.edt = kwargs.get('edt')
        self.tb_sql = "SELECT table_name FROM information_schema.tables WHERE table_schema='{db_name}'".format(
            db_name=kwargs.get('db', '')
        )
        self.area_sql = "SELECT area_id FROM {db_name}.%s LIMIT 1".format(
            db_name=kwargs.get('db', ''),
        )
        self.data_list = []
        self.w = Workbook()
        self.ws = self.w.active
        self.start = 2  # 合并单元格起始行
        self.end = 1  # 合并单元格结束行

        self.download_sum = 0
        self.push_sum = 0
        self.pub_sum = 0

        self.cwd = os.getcwd()
        self.file_path = os.path.join(os.path.join(self.cwd, 'files'),'统计{0:%Y-%m-%d}.xls'.format(datetime.now()))
        print(self.file_path)

    def get_statistic_data(self, date, table_with_area, serial_number):
        table_name = table_with_area['table_name']
        area_id = table_with_area['area_id']
        result = []

        # if area_id not in [14, 8, 3305]:
        today = date
        tomorrow = '{0:%Y-%m-%d}'.format(datetime.strptime(today, '%Y-%m-%d') + timedelta(days=1))
        download_sql = """SELECT area_id n, COUNT(id) c FROM {db_name}.{table_name}
        WHERE update_time BETWEEN '{today}' AND '{tomorrow}'
        GROUP BY area_id;""".format(**{
            'today': today,
            'tomorrow': tomorrow,
            'table_name': table_name,
            'db_name': self.db_name,
        })
        pub_sql = """SELECT area_id n, COUNT(id) c FROM {db_name}.{table_name}
        WHERE pub_time BETWEEN '{today}' AND '{tomorrow}'
        GROUP BY area_id;""".format(**{
            'today': today,
            'tomorrow': tomorrow,
            'table_name': table_name,
            'db_name': self.db_name,
        })
        push_sql = """SELECT area_id n, sum(count) FROM {db_name}.statistical
        WHERE date_format(push_time,'%Y-%m-%d')='{cdt}' AND area_id={area_id}
        """.format(**{
            'db_name': self.db_name,
            'area_id': area_id,
            'cdt': date,
        })
        # print(download_sql + '\n')
        # print(pub_sql + '\n')
        # print(push_sql + '\n')
        download_data = self.fetch_all(download_sql)  # [(area_id, n),]
        pub_data = self.fetch_all(pub_sql)
        push_data = self.fetch_all(push_sql)

        try:
            download_n = int(download_data[0][1])
        except Exception as e:
            download_n = 0
        try:
            pub_n = int(pub_data[0][1])
        except Exception as e:
            pub_n = 0
        try:
            push_n = int(push_data[0][1])
        except Exception as e:
            push_n = 0

        result.append(
            [date, serial_number, area_id, pub_n, download_n, push_n])

        self.data_list.extend([[x[0], x[1], self.area_map.get(str(x[2]), x[2]), x[3], x[4], x[5]] for x in result])

    @property
    def table_info(self):
        """
        表名对应的area_id
        Returns:

        """
        com = re.compile('(\d+)')
        ta = []
        tables = [n[0] for n in self.fetch_all(self.tb_sql) if n and n[0].startswith('notices')]
        for t in tables:
            area_ids = com.findall(t)
            # area_ids = self.fetch_all(self.area_sql % t)
            if area_ids:
                ta.append({
                    'table_name': t,
                    'area_id': area_ids[0]
                })
        return ta

    @staticmethod
    def get_date_list(sdt, edt):
        date_list = []
        if sdt == edt:
            date_list = ['{0:%Y-%m-%d}'.format(sdt)]
        if edt > sdt:
            days = (edt - sdt).days
            for _ in range(0, days + 1):
                date_list.append('{0:%Y-%m-%d}'.format(sdt))
                sdt += timedelta(days=1)
        return date_list

    def output(self, **kwargs):
        tts = ['日期', '序号', '网站名称', '站点发布数', '采集数量', '推送数量', '发布数量', '待发布数量']
        sdt = kwargs.get('sdt')
        edt = kwargs.get('edt')

        try:
            sdt = datetime.strptime(sdt, '%Y-%m-%d')
            edt = datetime.strptime(edt, '%Y-%m-%d')
        except ValueError:
            pass
        else:
            # datetime list
            date_list = ReportOutput.get_date_list(sdt, edt)

            # get all notices table_name
            table_info = self.table_info

            # DATA
            for date in date_list:
                print(date)
                self.data_list = []

                gevent.joinall([gevent.spawn(self.get_statistic_data, date, table_with_area, serial_number + 1) for
                                serial_number, table_with_area in enumerate(table_info)])

                try:
                    self.to_excel(tts, self.data_list)
                except Exception as e:
                    print(e)
                    self.msg = e
            # 合计
            for n, v in enumerate(['合计', '', '', self.pub_sum, self.download_sum, self.push_sum, '']):
                self.ws.cell(row=self.end + 1, column=n + 1, value=v)
            self.ws['A{0}'.format(self.end + 1)].border = self.border
            self.w.save(self.file_path)

    def to_excel(self, tts, data_list):
        """
        输出excel报表
        @result: [
            ['2021-04-16','医疗采购', 575， 1],
            ['2021-04-16','工程建设', 138， 2],
            ['2021-04-17','测试1', 157，0],
            ['2021-04-17','测试2', 3，0],
        ]
        params:
            @tts: 表头列表
            @data_list: 表体数据
            @date_list: 合并时间
        """
        for n, tt in enumerate(tts):
            self.ws.cell(row=1, column=n + 1, value=tt)

        for n, dl in enumerate(data_list):
            self.ws.append(dl)
            self.end += 1

            self.pub_sum += dl[3]
            self.download_sum += dl[4]
            self.push_sum += dl[5]

        # 合并
        self.ws.merge_cells('A{start}:A{end}'.format(start=self.start, end=self.end))
        # 居中
        self.ws['A{start}'.format(start=self.start)].alignment = styles.Alignment(
            horizontal="center", vertical="center"
        )
        # 样式
        for col in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']:
            self.ws[col].border = self.border
        for row in self.ws['A']:
            row.border = self.border

        self.w.save(self.file_path)
        self.start = self.end + 1


if __name__ == '__main__':
    data = {
        'host': '114.67.84.76',
        'user': 'root',
        'password': 'Ly3sa%@D0$pJt0y6',
        # 'db': 'test2_data_collection',
        'db': 'data_collection',
    }
    rpt = ReportOutput(**data)
    start_time = datetime.now()
    rpt.output(sdt='2021-05-28', edt='2021-05-29')
    print((datetime.now() - start_time).total_seconds())
