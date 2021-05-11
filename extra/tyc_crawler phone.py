"""
@file          :tyc_crawler phone.py
@description   :天眼查手机端根据关键词查找企业信息
@date          :2021/05/09 19:51:53
@author        :miaokela
@version       :1.0
"""

import requests
import gevent
import re
from lxml import etree
from functools import wraps
from datetime import datetime
from openpyxl import Workbook, styles
from openpyxl.styles import Border, Side


def time_count(func):
    @wraps(func)
    def inner(self):
        start_time = datetime.now()
        func(self)
        # 统计执行时间
        print((datetime.now() - start_time).total_seconds())

    return inner


class TYCCrawler(object):
    """
    天眼查站点信息抓取
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.72 Safari/537.36'
    }
    domain = 'https://m.tianyancha.com/'
    search_url = 'https://m.tianyancha.com/search?key={key}'
    detail_xpath = '//div[@class="search-name"]/a/@href'
    detail_info = {}  # {'招天下': 'www.www.www'}
    basic_info_re = '法定代表人,.*?,(?P<法定代表人>.*?),.*?成立日期,(?P<成立日期>.*?),经营状态,(?P<经营状态>.*?),注册资本,(?P<注册资本>.*?),' + \
                    '实缴资本,(?P<实缴资本>.*?),统一社会信用代码,(?P<统一社会信用代码>.*?),工商注册号,(?P<工商注册号>.*?),组织机构代码,(?P<组织机构代码>.*?),' + \
                    '纳税人识别号,(?P<纳税人识别号>.*?),纳税人资质,(?P<纳税人资质>.*?),企业类型,(?P<企业类型>.*?),行业,(?P<行业>.*?),' + \
                    '营业期限,(?P<营业期限始>.*?)至(?P<营业期限末>.*?),人员规模,(?P<人员规模>.*?),参保人数,(?P<参保人数>.*?),' + \
                    '英文名称,(?P<英文名称>.*?),曾用名,(?P<曾用名>.*?),登记机关,(?P<登记机关>.*?),核准日期,(?P<核准日期>.*?),注册地址,(?P<注册地址>.*?),经营范围,(?P<经营范围>.*)'

    def __init__(self, file_name):
        self.msg = ''
        self.file_name = file_name
        self.basic_info = {}

    @staticmethod
    def remove_specific_element(content, ele_name, attr_name, attr_value, if_child=False, index=1, text='', **kwargs):
        """
        remove specific html element attribute from content
        params:
            @content: html文档
            @ele_name: 元素名称
            @attr_name: 元素属性名
            @attr_value: 元素属性值
            @index: 指定元素索引删除
            @text: 指定包含文本的子节点删除
            @kwargs: if_child 移除的是否子元素
                    child_attr 子元素名称
        """
        msg = ''
        try:
            doc = etree.HTML(content)
            els = doc.xpath('//{ele_name}'.format(**{
                'ele_name': ele_name,
            }))

            if attr_name:
                same = 1
                for el in els:
                    if attr_value in el.get(attr_name, ''):
                        if if_child:
                            child_attr = kwargs.get('child_attr')

                            # if not text:
                            child_els = el.xpath(
                                './/{child_attr}'.format(**{'child_attr': child_attr}))

                            if not text:
                                for n, child_el in enumerate(child_els):
                                    if n == index - 1:
                                        child_el.getparent().remove(child_el)
                                        break
                            else:
                                for child_el in child_els:
                                    child_el_text = ','.join(
                                        child_el.xpath('.//text()'))
                                    if text in child_el_text:
                                        child_el.getparent().remove(child_el)
                                        break
                        else:
                            if index:
                                if same == index:
                                    el.getparent().remove(el)
                                    break
                                same += 1
                            else:  # index = 0 删除所有匹配节点
                                el.getparent().remove(el)
                content = etree.tounicode(doc)
            else:  # 无属性元素 指定索引删除
                for n, el in enumerate(els):
                    if n + 1 == index:
                        el.getparent().remove(el)
        except Exception as e:
            msg = e

        return msg, content.replace('<html><body>', '').replace('</body></html>', '')

    @property
    def kw_list(self):
        return ['上海特速网络科技有限公司', '浙江招天下招投标交易平台有限公司', '爱奇艺', '百度', '腾讯', '小米']
        # return ['浙江招天下招投标交易平台有限公司']

    @classmethod
    def get_page(cls, url, method='GET', **kwargs):
        """获取网页html文档

        Args:
            url (str): 链接地址
            method (str, optional): 请求方法. Defaults to 'GET'.

        Returns:
            str: 响应结果
        """
        ret = ''
        headers = kwargs.get('headers', '')
        headers = headers if headers else cls.headers
        if method == 'GET':
            ret = requests.get(
                url=url, headers=headers).content.decode('utf-8')
        if method == 'POST':
            ret = requests.post(
                url=url, headers=headers, data=kwargs.get('data', {})
            ).content.decode('utf-8')
        return ret

    def search_task(self, kd):
        """搜索企业链接地址

        Args:
            kd (str): 企业关键字
        """
        c_url = self.search_url.format(key=kd)
        resp = TYCCrawler.get_page(c_url)
        try:
            doc = etree.HTML(resp)
            detail_urls = doc.xpath(self.detail_xpath)
            if detail_urls:
                detail_url = detail_urls[0]
                self.detail_info[kd] = detail_url
        except Exception as e:
            self.msg = 'error:{e}'.format(e=e)

    def parse_basic_info_task(self, name, url):
        """获取企业基本信息
            统一社会信用代码|企业名称|注册号|组织机构代码|纳税人识别号|所属行业|法定代表人|公司类型|
            成立日期|注册资本|实缴资本|核准日期|营业期限自|营业期限至|登记机关|登记状态|注册地址|经营范围|

            企业名称|法定代表人 单独获取
        Args:
            name (str): 企业名称
            url (str): 详情页地址
        """
        resp = TYCCrawler.get_page(url)
        try:
            doc = etree.HTML(resp)

            # 企业名称
            company_els = doc.xpath('//div[@class="inner-name inner-title"]/text()')
            company = company_els[0] if company_els else ''

            els = doc.xpath('//div[@class="content" and position()=1]')

            if els:
                doc_content = etree.tounicode(els[0])

                # _, content = TYCCrawler.remove_specific_element(
                #     doc_content, 'div', 'class', 'data-describe ', index=0)

                c_doc = etree.HTML(doc_content)

                c_els = c_doc.xpath('//div[@class="divide-content"]/*')
                t = []
                for el in c_els:
                    t.append(','.join(
                        el.xpath('.//text()')
                    ).strip().replace(' ', '').replace('\n', ''))

                data = ','.join(t)

                com = re.compile(self.basic_info_re)

                ret = [m.groupdict() for m in re.finditer(com, data)]
                if ret:
                    self.basic_info[name] = ret[-1]
                    self.basic_info[name]['企业名称'] = company

        except Exception as e:
            self.msg = 'error:{e}'.format(e=e)

    @staticmethod
    def to_excel(tts, data):
        """
        导出至excel
        Args:
            tts: 表头信息
            data: 采集的数据

        Returns:
            msg: 错误信息
        """
        msg = ''
        border_type = Side(border_style="medium", color='FF000000')
        border = Border(
            left=border_type,
            right=border_type,
            top=border_type,
            bottom=border_type,
            diagonal=border_type,
            diagonal_direction=0,
            outline=border_type,
            vertical=border_type,
            horizontal=border_type
        )
        try:
            w = Workbook()
            ws = w.active
            for n, tt in enumerate(tts):
                ws.cell(row=1, column=n + 1, value=tt)
            # 样式
            for col in [
                'A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1',
                'Q1', 'R1', 'S1'
            ]:
                ws[col].border = border
            # {
            # 上海特速: {'法定代表人': '郑鲁', '成立日期': '2005-10-08', '经营状态': '存续', '注册资本': '1200万人民币', '实缴资本': '500万人民币', '统一社会信用代码': '91310114781132988W',
            # '工商注册号': '310114001535646', '组织机构代码': '781132988', '纳税人识别号': '91310114781132988W', '纳税人资质': '增值税一般纳税人', '企业类型': '有限责任公司(自然人投资或控股)',
            # '行业': '科技推广和应用服务业', '营业期限始': '2005-10-08', '营业期限末': '2035-10-07', '人员规模': '小于50人', '参保人数': '15', '英文名称': 'ShanghaiTesuNetworkTechnologyCo.,Ltd.',
            #  '曾用名': '-', '登记机关': '嘉定区市场监管局', '核准日期': '2017-12-13',
            # '注册地址': '上海市嘉定区封周路655号14幢201室J2434', '经营范围': '网络科技（不得从事科技中介），从事网络系统、计算机、软件及辅助设备技术领域内的技术开发、技术转让、技术咨询、技术服务，计算机系统集成，计算机、软件及辅助设备、文具用品、办公用品的销售。【依法须经批准的项目，经相关部门批准后方可开展经营活动】,更多', '企业名称': '上海特速网络科技有限公司'}
            # }
            n = 1
            for c, info in data.items():
                ws.append(([
                    n, c, info['统一社会信用代码'], info['工商注册号'], info['组织机构代码'], info['纳税人识别号'],
                    info['行业'], info['法定代表人'], info['企业类型'], info['成立日期'],
                    info['注册资本'], info['实缴资本'], info['核准日期'], info['营业期限始'],
                    info['营业期限末'], info['登记机关'], info['经营状态'], info['注册地址'],
                    info['经营范围'],
                ]))
                n += 1
            w.save('企业信息{0:%Y-%m-%d}.xls'.format(datetime.now()))
        except Exception as e:
            print(e)
            msg = 'error:{e}'.format(e=e)

        return msg

    @time_count
    def run(self, output_type=''):
        # 搜索所有关键字
        gevent.joinall([gevent.spawn(self.search_task, kd) for kd in self.kw_list])
        # 详情页信息
        gevent.joinall([gevent.spawn(self.parse_basic_info_task, name, url) for name, url in self.detail_info.items()])

        # 以excel表格的方式导出
        # '序号','企业名称','统一社会信用代码','注册号','组织机构代码','纳税人识别号','所属行业','法定代表人','公司类型','成立日期',
        # '注册资本','实缴资本','核准日期','营业期限自','营业期限至','登记机关','登记状态','注册地址','经营范围'
        tts = ['序号', '企业名称', '统一社会信用代码', '注册号', '组织机构代码', '纳税人识别号', '所属行业', '法定代表人', '公司类型',
               '成立日期', '注册资本', '实缴资本', '核准日期', '营业期限自', '营业期限至', '登记机关', '登记状态', '注册地址', '经营范围']
        self.to_excel(tts, self.basic_info)

        print(self.basic_info)
        print('finish.')


if __name__ == '__main__':
    file_name = "./kw.xml"
    tyc = TYCCrawler(file_name)
    tyc.run()
